import csv
import io
import uuid
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from backend.config import DATA_DIR

UPLOAD_DIR = DATA_DIR / "uploads"
MAX_DOC_CHARS = 10_000
MAX_TOTAL_CHARS = 30_000


def save_file(account_id: int, filename: str, content: bytes) -> tuple[str, str]:
    """Save uploaded file to disk. Returns (stored_filename, relative_file_path)."""
    account_dir = UPLOAD_DIR / str(account_id)
    account_dir.mkdir(parents=True, exist_ok=True)

    stored_name = f"{uuid.uuid4().hex[:12]}_{filename}"
    file_path = account_dir / stored_name
    file_path.write_bytes(content)

    return stored_name, str(file_path.relative_to(DATA_DIR.parent))


def delete_file(file_path: str) -> None:
    """Delete a file from disk."""
    full_path = DATA_DIR.parent / file_path
    if full_path.exists():
        full_path.unlink()


def extract_text(filename: str, content: bytes) -> str:
    """Extract text from uploaded file content. Returns empty string for unsupported types."""
    lower = filename.lower()
    try:
        if lower.endswith((".txt", ".md")):
            return content.decode("utf-8", errors="replace")
        if lower.endswith(".csv"):
            return _extract_csv(content)
        if lower.endswith(".docx"):
            return _extract_docx(content)
        if lower.endswith(".pdf"):
            return _extract_pdf(content)
        if lower.endswith(".xlsx"):
            return _extract_xlsx(content)
    except Exception:
        return ""
    return ""


def _extract_docx(content: bytes) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(content), "r") as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
            root = tree.getroot()
    lines = []
    for p in root.findall(".//w:p", ns):
        texts = p.findall(".//w:t", ns)
        line = "".join(t.text for t in texts if t.text)
        if line.strip():
            lines.append(line.strip())
    return "\n".join(lines)


def _extract_pdf(content: bytes) -> str:
    import pdfplumber
    text_parts = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def _extract_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip(" |"):
                parts.append(row_text)
    return "\n".join(parts)


def _extract_csv(content: bytes) -> str:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def get_account_document_context(documents) -> str:
    """Build a context string from a list of AccountDocument objects (all types)."""
    if not documents:
        return ""

    parts = ["## Uploaded Documents"]
    total_chars = 0

    for doc in documents:
        if doc.extracted_text:
            text = doc.extracted_text
            if len(text) > MAX_DOC_CHARS:
                text = text[:MAX_DOC_CHARS] + "\n\n[Document truncated...]"
            if total_chars + len(text) > MAX_TOTAL_CHARS:
                parts.append(f"\n- {doc.original_filename} (text omitted, context limit reached)")
                continue
            parts.append(f"\n### {doc.original_filename}")
            parts.append(text)
            total_chars += len(text)
        else:
            parts.append(f"\n- {doc.original_filename} (binary file, no text extracted)")

    return "\n".join(parts)


def get_knowledge_context(documents) -> str:
    """Build context from knowledge base documents (BrainLift, Account BrainLift)."""
    knowledge_docs = [d for d in documents if getattr(d, 'doc_type', 'activity') == 'knowledge']
    if not knowledge_docs:
        return ""

    parts = ["## Account Knowledge Base"]
    total_chars = 0

    for doc in knowledge_docs:
        if doc.extracted_text:
            text = doc.extracted_text
            if len(text) > MAX_DOC_CHARS:
                text = text[:MAX_DOC_CHARS] + "\n\n[Document truncated...]"
            if total_chars + len(text) > MAX_TOTAL_CHARS:
                continue
            parts.append(f"\n### {doc.original_filename}")
            parts.append(text)
            total_chars += len(text)

    return "\n".join(parts)


def get_activity_context(documents) -> str:
    """Build context from activity documents (call transcripts, emails, meeting notes)."""
    activity_docs = [d for d in documents if getattr(d, 'doc_type', 'activity') == 'activity']
    if not activity_docs:
        return ""

    parts = ["## Recent Activity (calls, emails, notes)"]
    total_chars = 0

    for doc in activity_docs:
        if doc.extracted_text:
            text = doc.extracted_text
            if len(text) > MAX_DOC_CHARS:
                text = text[:MAX_DOC_CHARS] + "\n\n[Document truncated...]"
            if total_chars + len(text) > MAX_TOTAL_CHARS:
                continue
            parts.append(f"\n### {doc.original_filename}")
            parts.append(text)
            total_chars += len(text)

    return "\n".join(parts)
